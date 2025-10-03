from io import BytesIO
from django.template.loader import render_to_string
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from weasyprint import HTML


def export_reports_xlsx(queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Reports'
    headers = ['ID Number','Date','Type','Referred By','Sonologist','Total Ultra','Patient','Notes']
    ws.append(headers)
    for r in queryset:
        ws.append([
        r.id_number, r.date.isoformat(), r.type_of_usg, r.referred_by, r.sonologist, r.total_ultra, r.patient_name, (r.notes or '')
        ])
    # autosize columns
    for i, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = min(50, max(len(str(cell.value or '')) for cell in col) + 2)
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    response = HttpResponse(stream.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=reports.xlsx'
    return response


def export_reports_pdf(queryset, request=None):
    html_string = render_to_string('reports/report_pdf.html', {'reports': queryset})
    html = HTML(string=html_string, base_url=request.build_absolute_uri('/') if request else None)
    pdf = html.write_pdf()
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reports.pdf"'
    return response